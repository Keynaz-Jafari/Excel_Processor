from openpyxl import load_workbook

file_path = 'knoush.xlsx'
workbook = load_workbook(filename=file_path)
sheet = workbook.active

name_counts = {}
night_shift_counts = {}

# Iterate through rows to collect all names and their counts
for row_index in range(3, sheet.max_row + 1):
    for cell in sheet[row_index][2:]:  # Start from column index 2 (C column)
        name = cell.value
        if name:
            # Get the drugstore name based on the column
            if 3 <= cell.column <= 4:
                drugstore_name = sheet.cell(row=1, column=3).value
            elif 5 <= cell.column <= 7:
                drugstore_name = sheet.cell(row=1, column=5).value
            elif 8 <= cell.column <= 9:
                drugstore_name = sheet.cell(row=1, column=8).value
            elif 10 <= cell.column <= 12:
                drugstore_name = sheet.cell(row=1, column=10).value
            elif 13 <= cell.column <= 15:
                drugstore_name = sheet.cell(row=1, column=13).value
            elif 16 <= cell.column <= 18:
                drugstore_name = sheet.cell(row=1, column=16).value
            elif cell.column == 19:
                drugstore_name = sheet.cell(row=1, column=19).value
            elif cell.column == 20:
                drugstore_name = sheet.cell(row=1, column=20).value
            else:
                drugstore_name = None

            # Increment the count for the current name in the corresponding drugstore
            if name not in name_counts:
                name_counts[name] = {drugstore_name: 1}
            else:
                if drugstore_name not in name_counts[name]:
                    name_counts[name][drugstore_name] = 1
                else:
                    name_counts[name][drugstore_name] += 1
            
            # Check if the column is L, O, or R for night shifts
            if cell.column in [12, 15, 18]:
                if name not in night_shift_counts:
                    night_shift_counts[name] = 1
                else:
                    night_shift_counts[name] += 1

report_file_path = 'name_counts_report.txt'
with open(report_file_path, 'w') as report_file:
    num = 1
    for name, counts in name_counts.items():
        report_file.write(str(num) + '. ' + f"Name: {name}\n")
        for drugstore, count in counts.items():
            # Include "(night shift)" label for L, O, R columns
            night_shift_label = " (night shift)" if drugstore and drugstore[-1] in ['L', 'O', 'R'] else ""
            report_file.write(f"{drugstore}\t{count}{night_shift_label}\n")
        report_file.write('\n')
        num += 1
    
    # Write total night shift counts at the end
    report_file.write("Total Night Shift Counts:\n")
    for name, count in night_shift_counts.items():
        report_file.write(f"{name}\t{count}\n")

print(f"Name counts report saved as {report_file_path}")
