import pandas as pd
from openpyxl import load_workbook
import random

file_path = 'Farvardin403.xlsx'
workbook = load_workbook(filename=file_path)

sheet = workbook.active

for row_index, row in enumerate(sheet.iter_rows(), start=1):
    # a set to store names already written in the current row
    names_set = set()

    # variables to store names from the previous row
    prev_names_col12 = set()
    prev_names_col13 = set()
    prev_names_col15 = set()
    prev_names_col16 = set()

    # Extract names from the previous row for columns 12, 13, 15, and 16
    if row_index > 1:
        prev_row = sheet[row_index - 1]
        prev_names_col12 = {cell.value for cell in prev_row[11:12]}
        prev_names_col13 = {cell.value for cell in prev_row[12:13]}
        prev_names_col15 = {cell.value for cell in prev_row[14:15]}
        prev_names_col16 = {cell.value for cell in prev_row[15:16]}

    for cell_index, cell in enumerate(row, start=1):
        comment = cell.comment
        if comment:
            # Extract comments
            comment_text = comment.text

            lines = comment_text.split('\n')

            # Extract names from comments
            extracted_names = []
            for line in lines:
                name = line.strip().split('-')[0].strip()  # Extract the name part before the '-'
                if name and name not in extracted_names:
                    extracted_names.append(name)

            random.shuffle(extracted_names)

            # Find the first non-repeated name
            name_to_write = None
            for name in extracted_names:
                if name not in names_set:
                    name_to_write = name
                    break

            #write the name to the cell if the name is not repeated 
            if name_to_write:
                cell.value = name_to_write
                names_set.add(name_to_write)

            if cell_index == 10 and name_to_write in (prev_names_col12 | prev_names_col13):
                # If the name is repeated in the previous row of column 12 or 13, select another name
                other_names = names_set - {name_to_write}
                if other_names:
                    new_name = random.choice(list(other_names))
                    cell.value = new_name
                    names_set.add(new_name)
            elif cell_index == 13 and name_to_write in prev_names_col15:
                # If the name is repeated in the previous row of column 15, select another name
                other_names = names_set - {name_to_write}
                if other_names:
                    new_name = random.choice(list(other_names))
                    cell.value = new_name
                    names_set.add(new_name)
            elif cell_index == 15 and name_to_write in prev_names_col16:
                # If the name is repeated in the previous row of column 16, select another name
                other_names = names_set - {name_to_write}
                if other_names:
                    new_name = random.choice(list(other_names))
                    cell.value = new_name
                    names_set.add(new_name)


output_excel_file = 'Farvardin403_with_values.xlsx'
workbook.save(output_excel_file)

print(f"Modified Excel file saved as {output_excel_file}")
